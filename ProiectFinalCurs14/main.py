import openpyxl
from openpyxl.chart import (PieChart, Reference)
from tkinter import *
from win32com import client


root = Tk()
root.geometry('400x400')
root.title('Test Cases Parser')
Label(root, text='Test Cases Parser', font=' arial 25 bold').pack()

tester_label = Label(root, text='Tester Name: ', font='arial 10 bold').pack()
tester_str = StringVar()
Entry(root, textvariable=tester_str).pack()

path=r"D:\Anca\Git\Proiect Final\Proiect-Final\Test_Case_Anca_Dinu.xlsx"
path_PDF=r"D:\Anca\Git\Proiect Final\Proiect-Final\Test_Case_Anca_Dinu.pdf"

values=[0,0]
passString = 'PASS'
failString = 'FAIL'


def compareValues():
    wb = openpyxl.load_workbook(path, read_only=False)
    testCasesSheet = wb['test cases']
    for row in range(11, int(testCasesSheet.max_row + 1)):
        if testCasesSheet.cell(row=row, column=7).value == passString.upper():
            values[0] = values[0]+1

        elif testCasesSheet.cell(row=row, column=7).value == failString.upper():
            values[1]=values[1]+1

    totalTestCases_counter = values[1]+values[0]
    print("Total teste Fail: ", values[1])
    print("Total teste Pass: ", values[0])
    print('Total Teste: ', totalTestCases_counter)

def generateReport():
    wb = openpyxl.load_workbook(path, read_only=False)

    first_sheet = wb['test cases']
    tester = first_sheet['E1'].value

    try:
        reportSheet = wb['Report']
    except:
        wb.create_sheet('Report')
        reportSheet = wb.get_sheet_by_name('Report')

    reportSheet['A1'] = 'TesterID: '
    reportSheet['B1'] = tester
    reportSheet['A2'] = 'Failed test cases'
    reportSheet['B2'] = values[1]
    reportSheet['A3'] = 'Passed test cases'
    reportSheet['B3'] = values[0]
    reportSheet['A4'] = 'Total number of test cases'
    reportSheet['B4'] = values[0] + values[1]
    wb.save(path)
    createChart()

    excel = client.Dispatch("Excel.Application")

    sheets = excel.Workbooks.Open(path)
    work_sheets = sheets.Worksheets[2]
    work_sheets.ExportAsFixedFormat(0, path_PDF)

def createChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb['Report']
    pie = PieChart()

    labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(sheet, min_col=2, min_row=2, max_row=3)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Test Cases"

    pie.width = 14
    pie.height = 7
    s = pie.series[0]

    s.graphicalProperties.line.solidFill = "00000"

    sheet.add_chart(pie, 'A6')
    wb.save(path)



def buttonPressed():

    compareValues()
    generateReport()




Button(root, text='Generate Report', command=buttonPressed).pack(pady=10)

root.mainloop()




